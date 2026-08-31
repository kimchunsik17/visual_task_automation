"""documents/xlsx_builder.py — DocumentSpec(5블록)을 .xlsx 로 렌더한다.

문서 흐름을 시트 위 행 흐름으로 옮긴다:
  heading    → 큰 글씨·굵은 행 (level 로 크기 차등)
  paragraph  → 일반 행 (줄바꿈 허용, A열 넓게)
  table      → 헤더(굵게+회색 배경) + 격자 테두리
  image      → 셀 앵커에 삽입
  page_break → **새 시트** (엑셀에서 쪽의 자연스러운 대응물)

hwpx/docx builder 와 같은 실패 계약(SpecError, 조용히 빠뜨리지 않음)을 지킨다.
"""

from __future__ import annotations

import io
import os
import re
from typing import Any, Dict, Optional

from .hwpx.builder import (
    IMAGE_FORMATS, MAX_BLOCKS, MAX_TABLE_CELLS, SUPPORTED_BLOCKS, ImageLoader, SpecError,
    UnsupportedFeature,
)

_HEADING_PT = {1: 16, 2: 13, 3: 11}
_MAX_SHEET_TITLE = 31  # 엑셀 제한


def _sheet_title(base: str, index: int) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", " ", base or "Sheet").strip() or "Sheet"
    suffix = f" ({index})" if index > 1 else ""
    return (cleaned[: _MAX_SHEET_TITLE - len(suffix)] + suffix) or f"Sheet{index}"


def build(spec: Dict[str, Any], output_path: str,
          image_loader: Optional[ImageLoader] = None) -> Dict[str, Any]:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XlsxImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if not isinstance(spec, dict):
        raise SpecError("DocumentSpec 은 객체여야 합니다.")
    blocks = spec.get("blocks")
    if not isinstance(blocks, list) or not blocks:
        raise SpecError("blocks 가 비어 있습니다.")
    if len(blocks) > MAX_BLOCKS:
        raise SpecError(f"블록이 너무 많습니다({len(blocks)}개, 상한 {MAX_BLOCKS}개).")

    title = str(spec.get("title") or "문서")
    workbook = Workbook()
    sheet_index = 1
    sheet = workbook.active
    sheet.title = _sheet_title(title, sheet_index)
    cursor = 1
    max_width_cols = 1

    thin = Side(style="thin", color="D0D5DD")
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EEF2F7")
    wrap = Alignment(wrap_text=True, vertical="top")

    for index, block in enumerate(blocks):
        if not isinstance(block, dict):
            raise SpecError(f"blocks[{index}] 는 객체여야 합니다.")
        btype = block.get("type")
        if btype not in SUPPORTED_BLOCKS:
            raise UnsupportedFeature(
                f"blocks[{index}].type {btype!r} 는 지원하지 않습니다({', '.join(SUPPORTED_BLOCKS)}).")

        if btype == "heading":
            level = block.get("level", 1)
            if level not in _HEADING_PT:
                raise SpecError(f"heading.level 은 {sorted(_HEADING_PT)} 중 하나여야 합니다.")
            text = block.get("text")
            if not isinstance(text, str) or not text:
                raise SpecError("heading.text 가 비어 있습니다.")
            cell = sheet.cell(row=cursor, column=1, value=text)
            cell.font = Font(bold=True, size=_HEADING_PT[level])
            cursor += 2  # 제목 아래 한 행 띄움

        elif btype == "paragraph":
            text = block.get("text")
            if text is None:
                text = ""
            if not isinstance(text, str):
                raise SpecError("paragraph.text 는 문자열이어야 합니다.")
            cell = sheet.cell(row=cursor, column=1, value=text)
            cell.alignment = wrap
            cursor += 1

        elif btype == "table":
            columns = block.get("columns")
            rows = block.get("rows")
            if not isinstance(columns, list) or not columns:
                raise SpecError("table.columns 는 비어 있지 않은 배열이어야 합니다.")
            if not isinstance(rows, list):
                raise SpecError("table.rows 는 배열이어야 합니다.")
            width = len(columns)
            if width * (len(rows) + 1) > MAX_TABLE_CELLS:
                raise SpecError(f"표가 너무 큽니다(상한 {MAX_TABLE_CELLS}칸).")
            max_width_cols = max(max_width_cols, width)
            for column_index, header in enumerate(columns, start=1):
                cell = sheet.cell(row=cursor, column=column_index,
                                  value="" if header is None else str(header))
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.border = grid
            cursor += 1
            for row_offset, row in enumerate(rows):
                if not isinstance(row, list) or len(row) != width:
                    raise SpecError(f"table.rows[{row_offset}] 의 칸 수가 columns 와 다릅니다.")
                for column_index, value in enumerate(row, start=1):
                    cell = sheet.cell(row=cursor, column=column_index,
                                      value="" if value is None else str(value))
                    cell.border = grid
                    cell.alignment = wrap
                cursor += 1
            cursor += 1  # 표 아래 한 행 띄움

        elif btype == "image":
            artifact_id = block.get("artifactId")
            if not artifact_id:
                raise SpecError("image 블록은 artifactId 로만 지정할 수 있습니다(경로·URL 은 받지 않습니다).")
            if image_loader is None:
                raise SpecError("이 실행에서는 이미지를 넣을 수 없습니다(Artifact 해석기가 없습니다).")
            data, image_format = image_loader(str(artifact_id))
            image_format = str(image_format or "").lower().lstrip(".")
            if image_format not in IMAGE_FORMATS:
                raise SpecError(f"지원하지 않는 이미지 형식입니다: {image_format or '(알 수 없음)'}")
            image = XlsxImage(io.BytesIO(data))
            sheet.add_image(image, f"A{cursor}")
            cursor += max(1, int(image.height / 19) + 1)  # 기본 행 높이 ≈19px 기준 자리 확보

        else:  # page_break → 새 시트
            sheet_index += 1
            sheet = workbook.create_sheet(_sheet_title(title, sheet_index))
            cursor = 1

    # 읽기 편의: 본문 열 폭
    for ws in workbook.worksheets:
        ws.column_dimensions["A"].width = 46
        for col in range(2, max_width_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 24

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    workbook.save(output_path)
    return {"path": output_path, "blocks": len(blocks), "sheets": sheet_index}
